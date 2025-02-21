import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import PatternFill

fill_color = PatternFill(
    start_color="FFFF00",
    end_color="FFFF00",
    fill_type="solid"
    )

def insert_to_file(filename: str, data):
    workbook = load_workbook(filename)
    df_new = pd.DataFrame(data)

    rows = dataframe_to_rows(df_new, index=False, header=False)
    for row in rows:
        workbook.active.append(row)

    workbook.save(filename)


def add_detail_name(detail_name, file_name):
    workbook = load_workbook(file_name)
    sheet = workbook.active
    start_row = sheet.max_row + 1
    sheet.cell(row=start_row, column=1).value = detail_name
    sheet.cell(row=start_row, column=1).fill = fill_color
    workbook.save(file_name)


def fill_the_row(file_name):
    workbook = load_workbook(file_name)
    sheet = workbook.active
    start_row = sheet.max_row + 1
    for col in range(1, sheet.max_column + 1):
        cell = sheet.cell(row=start_row, column=col)
        cell.fill = fill_color
    workbook.save(file_name)

